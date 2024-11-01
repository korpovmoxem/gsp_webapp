from datetime import datetime
import base64
import os

from fastapi import Request
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from tempfile import NamedTemporaryFile
import smtplib
from email.mime.text import MIMEText
from email.utils import make_msgid
from email.utils import formataddr

from tools.tools import fast_bitrix_slow as fast_bitrix

USER_LIST = {
    '3044': 'ChuprinAVI@GSPROM.RU',
    '3035': 'ShkaretnykhAA@GSPROM.RU',
    '3031': 'LaubganDA@GSPROM.RU',
    '3045': 'SavichevKE@GSPROM.RU',
    '3027': 'SudnikovEV@GSPROM.RU',
    '2964': 'VirokhobskiiAA@GSPROM.RU',
    '3504': 'BalabaAD@GSPROM.RU',
    '3685': 'KozlovVV@GSPROM.RU',
    '3195': 'KerroSSE@GSPROM.RU',
    '2861': 'BelovaAOl@GSPROM.RU',
    '2996': 'KolodiazhnyiGIu@GSPROM.RU',
    '3030': 'ZakharovRA@GSPROM.RU',
    '3040': 'ZhukovaVI@GSPROM.RU',
    '2710': 'VolkovAV@GSPROM.RU',
    '3041': 'VasilchikovIS@GSPROM.RU',
    '2711': 'KhoviakovaOA@GSPROM.RU',
    '3032': 'MukhametshinaAA@GSPROM.RU',
    '3621': 'EremeevaPIU@GSPROM.RU',
    '3043': 'GurakIuV@GSPROM.RU',
    '3036': 'RomanenkoEN@GSPROM.RU',
    '3039': 'DzetovetskiiIV@GSPROM.RU',
    '3023': 'GribachevSPE@GSPROM.RU'
}


def get_element_value(field_name: str, fields_info: dict, element_info: dict) -> str:
    try:
        value = list(element_info[fields_info[field_name]['FIELD_ID']].values())[0]
        if '|RUB' in value:
            value = value.replace('|RUB', '')
        return value
    except KeyError:
        return '0'
    except AttributeError:
        return element_info[fields_info[field_name]['FIELD_ID']]


def get_indicator_color(main_value: int, other_value: int) -> str:
    #try:
        #percent_value = int(int(other_value) * 100 / int(main_value))
    #except:
        #return 'ff0000'
    try:
        other_value = int(other_value)
    except:
        return 'ff0000'
    if other_value < 50:
        return 'ff0000'
    elif other_value < 80:
        return 'ffbf00'
    else:
        return '00b04f'


def get_user_folder_id(fast_bitrix, user_id: str, folder_name: str = 'РАС') -> str:
    storage_info = fast_bitrix.get_all('disk.storage.getlist', {
        'filter': {
            'ENTITY_TYPE': 'user',
            'ENTITY_ID': user_id,
        }
    })

    storage_id = storage_info[0]['ID']
    storage_folders = fast_bitrix.get_all('disk.storage.getchildren', {
        'id': storage_id
    })
    report_folder = list(filter(lambda x: x['NAME'] == folder_name, storage_folders))
    if report_folder:
        return report_folder[0]['ID']

    new_folder = fast_bitrix.call('disk.storage.addfolder', {
        'id': storage_id,
        'data': {
            'NAME': folder_name
        }
    })
    return new_folder['ID']


def get_report_months(year_display_values, month_display_values):
    months = {
        1: 'ЯНВАРЬ',
        2: 'ФЕВРАЛЬ',
        3: 'МАРТ',
        4: 'АПРЕЛЬ',
        5: 'МАЙ',
        6: 'ИЮНЬ',
        7: 'ИЮЛЬ',
        8: 'АВГУСТ',
        9: 'СЕНТЯБРЬ',
        10: 'ОКТЯБРЬ',
        11: 'НОЯБРЬ',
        12: 'ДЕКАБРЬ'
    }
    result = list()
    current_year = datetime.now().year
    current_month = datetime.now().month
    for i in range(1):
        current_month -= 1
        if current_month == 0:
            current_month = 12
            current_year -= 1
        result.insert(0, ((list(filter(lambda x: x[1] == months[current_month], list(month_display_values.items())))[0][0], list(filter(lambda x: x[1] == str(current_year), list(year_display_values.items())))[0][0])))
    result.append(((list(filter(lambda x: x[1] == months[datetime.now().month], list(month_display_values.items())))[0][0], list(filter(lambda x: x[1] == str(datetime.now().year), list(year_display_values.items())))[0][0])))
    current_year = datetime.now().year
    current_month = datetime.now().month
    for i in range(2):
        current_month += 1
        if current_month == 13:
            current_month = 1
            current_year += 1
        result.append(((list(filter(lambda x: x[1] == months[current_month], list(month_display_values.items())))[0][0], list(filter(lambda x: x[1] == str(current_year), list(year_display_values.items())))[0][0])))
    return result


def get_element_row_value(elem, responsible, elem_type, month_name, year, fields_info):
    row_values = [
            elem['NAME'],
            f"{responsible['LAST_NAME']} {responsible['NAME']}",
            month_name,
            year,
            get_element_value('План ремонта ТС, шт', fields_info, elem),
            get_element_value('Перенос с предыдущего месяца', fields_info, elem),
            get_element_value('ИТОГО план ремонтов ТС на 1 число, шт', fields_info, elem),
            get_element_value('Дефектовка ТС, шт', fields_info, elem),
            get_element_value('Дефектовка, %', fields_info, elem),
            get_element_value('Заявка на ремонт ТС, шт', fields_info, elem),
            get_element_value('Заявка на ремонт, %', fields_info, elem),
            get_element_value('Закупка ЗЧ для ТС, шт', fields_info, elem),
            get_element_value('Закупка ЗЧ для ТС, %', fields_info, elem),
            get_element_value('Оплата ЗЧ для ТС, шт', fields_info, elem),
            get_element_value('Оплата ЗЧ для ТС, %', fields_info, elem),
            get_element_value('Поставка ЗЧ для ТС, шт', fields_info, elem),
            get_element_value('Поставка ЗЧ для ТС, %', fields_info, elem),
            get_element_value('Ремонт ТС, шт', fields_info, elem),
            get_element_value('Ремонт ТС факт, %', fields_info, elem),
            f"{round(float(get_element_value('ОБС с учетом КЗ, План', fields_info, elem))):_}".replace('_', ' '),
            f"{round(float(get_element_value('ОБС с учетом КЗ, Факт', fields_info, elem))):_}".replace('_', ' '),
            f"{round(float(get_element_value('ОперШтаб, План', fields_info, elem))):_}".replace('_', ' '),
            f"{round(float(get_element_value('ОперШтаб, Факт', fields_info, elem))):_}".replace('_', ' '),
            f"{round(float(get_element_value('Страховой запас, План', fields_info, elem))):_}".replace('_', ' '),
            f"{round(float(get_element_value('Страховой запас, Факт', fields_info, elem))):_}".replace('_', ' '),
            f"{round(float(get_element_value('Мониторинговый счет, Факт', fields_info, elem))):_}".replace('_', ' '),
            f"{round(float(get_element_value('Мониторинговый счет, Остаток на р/сч', fields_info, elem))):_}".replace('_', ' '),
            f"{round(float(get_element_value('Итого ФАКТ по хабу, руб.', fields_info, elem))):_}".replace('_', ' '),
        ]
    if elem_type == 'summary_month':
        row_values[1] = ''
        row_values[3] = ''
    return row_values


def excel_list_28():
    """
    Формирование отчета в формате .xlsx из данных списка и загрузка файла на диск.
    Пользователю, запустившему процесс, приходит ссылка на отчет в уведомлениях.

    :param fast_bitrix: Асинхронный экземпляр класса Bitrix библиотеки fast_bitrix24
    :param params: Параметры запроса:
                    list_id: ID списка
                    year: Год отчета
    """

    report_titles = [
        [
            'Проект', 'Ответственный', 'Месяц', 'Год', 'План ремонта ТС, шт', 'Перенос с предыдущего месяца',
            'ИТОГО план ремонтов ТС на 1 число, шт',
            'Дефектовка ТС', '', 'Заявка на ремонт ТС', '', 'Закупка ЗЧ для ТС', '', 'Оплата ЗЧ для ТС', '',
            'Поставка ЗЧ для ТС', '',
            'Ремонт ТС', '', 'ОБС с учетом КЗ, руб.', '', 'ОперШтаб, руб.', '', 'Страховой запас', '',
            'Мониторинговый счет, Факт',
            'Мониторинговый счет, Остаток на р/сч', 'Итого ФАКТ по хабу, руб'
        ],
        [
            '', '', '', '', '', '', '', 'шт.', '%', 'шт.', '%', 'шт.', '%', 'шт.', '%', 'шт.', '%', 'шт.', '%',
            'План', 'Факт', 'План', 'Факт', 'План', 'Факт',
        ]
    ]
    first_row_title = ['Свод по ремонтам технических ресурсов в разрезе Хаба'] + [''] * (len(report_titles[0]) - 3) + [
        datetime.now().strftime('%d.%m.%Y')]
    report_titles.insert(0, first_row_title)

    book = openpyxl.Workbook()
    sheet = book.active

    for row in report_titles:
        sheet.append(row)

    sheet.merge_cells('A2:A3')  # Проект
    sheet.merge_cells('B2:B3')  # Ответственный
    sheet.merge_cells('C2:C3')  # Месяц
    sheet.merge_cells('D2:D3')  # Год
    sheet.merge_cells('E2:E3')  # План ремонта ТС, шт
    sheet.merge_cells('F2:F3')  # Перенос с предыдущего месяца
    sheet.merge_cells('G2:G3')  # ИТОГО план ремонтов ТС на 1 число, шт
    sheet.merge_cells('A1:Z1')  # Свод по ремонтам технических ресурсов в разрезе Хаба
    sheet.merge_cells('AA1:AB1')  # datetime.now()
    sheet.merge_cells('H2:I2')  # Дефектовка ТС
    sheet.merge_cells('J2:K2')  # Заявка на ремонт ТС
    sheet.merge_cells('L2:M2')  # Закупка ЗЧ для ТС
    sheet.merge_cells('N2:O2')  # Оплата ЗЧ для ТС
    sheet.merge_cells('P2:Q2')  # Поставка ЗЧ для ТС
    sheet.merge_cells('R2:S2')  # Ремонт ТС
    sheet.merge_cells('T2:U2')  # ОБС с учетом КЗ, руб.
    sheet.merge_cells('V2:W2')  # ОперШтаб, руб.
    sheet.merge_cells('X2:Y2')  # Страховой запас
    sheet.merge_cells('Z2:Z3')  # ОБС с учетом КЗ, руб.
    sheet.merge_cells('AA2:AA3')  # ОперШтаб, руб.
    sheet.merge_cells('AB2:AB3')  # Страховой запас

    fields_raw = fast_bitrix.get_all('lists.field.get', {
        'IBLOCK_TYPE_ID': 'lists',
        'IBLOCK_ID': '28',
    })

    fields_info = dict()
    for key in fields_raw.keys():
        fields_info[fields_raw[key]['NAME']] = fields_raw[key]

    date_elements = list(get_report_months(fields_info['Год']['DISPLAY_VALUES_FORM'], fields_info['Месяц']['DISPLAY_VALUES_FORM']))

    '''
    elements = await fast_bitrix.get_all('lists.element.get', {
        'IBLOCK_TYPE_ID': 'lists',
        'IBLOCK_ID': '28',
        'FILTER': {
            fields_info['Год']['FIELD_ID']: year_code,
        }
    })
    '''
    responsible_field_id = fields_info['Ответственный']['FIELD_ID']
    for date_tuple in date_elements:
        month_elements = fast_bitrix.get_all('lists.element.get', {
            'IBLOCK_TYPE_ID': 'lists',
            'IBLOCK_ID': '28',
            'FILTER': {
                fields_info['Год']['FIELD_ID']: date_tuple[1],
                fields_info['Месяц']['FIELD_ID']: date_tuple[0],
            }
        })
        users_info = fast_bitrix.get_all('user.get', {
            'FILTER': {
                'ID': list(map(lambda x: get_element_value('Ответственный', fields_info, x),
                               filter(lambda y: responsible_field_id in y, month_elements)))
            }
        })
        month_field_id = fields_info['Месяц']['FIELD_ID']
        month_name = list(filter(lambda x: x[0] == date_tuple[0], fields_info['Месяц']['DISPLAY_VALUES_FORM'].items()))[0][1]
        year_name = list(filter(lambda x: x[0] == date_tuple[1], fields_info['Год']['DISPLAY_VALUES_FORM'].items()))[0][1]
        if not month_elements:
            continue
        for elem_index, elem in enumerate(sorted(filter(lambda x: fields_info['Ответственный']['FIELD_ID'] in x, month_elements), key=lambda x: x['NAME'])):
            responsible = list(filter(lambda x: x['ID'] == get_element_value('Ответственный', fields_info, elem), users_info))[0]
            sheet.append(get_element_row_value(elem, responsible, 'row', month_name, year_name, fields_info))
        summary_month_element = list(filter(lambda element: get_element_value('Месяц', fields_info, element) == date_tuple[0] and get_element_value('Год', fields_info, element) == date_tuple[1] and 'Итого за месяц' in element['NAME'], month_elements))
        if not summary_month_element:
            sheet.append([])
            continue
        sheet.append(get_element_row_value(summary_month_element[0], responsible, 'summary_month', month_name, year_name, fields_info))
        for col_index, _ in enumerate(next(sheet.rows), 1):
            cell = sheet.cell(len(list(sheet.rows)), col_index)
            cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
            cell.font = Font(bold=True)

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row_index, row in enumerate(sheet.rows, 1):
        for cell_index, cell in enumerate(row, 1):
            if cell_index in (1, 2) and row_index not in (1,):
                cell.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            cell.border = thin_border

            if row_index in (1, 2, 3):
                cell.font = Font(bold=True)

            if get_column_letter(cell_index) in ('I', 'K', 'M', 'O', 'Q', 'S') and row_index >= 3:
                cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                if cell.value and '%' not in cell.value:
                    cell_color = get_indicator_color(sheet.cell(row_index, cell_index).value, cell.value)
                    cell.fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type='solid')
                    cell.value = cell.value + '%'

    for column_index, _ in enumerate(sheet.columns, 1):
        col_letter = get_column_letter(column_index)
        sheet.column_dimensions[col_letter].width = 11

    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 30

    for row_index, _ in enumerate(sheet.rows, 1):
        height = 18
        if row_index in (1, 2, 3):
            height = 30
        sheet.row_dimensions[row_index].height = height

    sheet = book.create_sheet('Комплектация хабов тех.ресурсами')
    fields_raw = fast_bitrix.get_all('lists.field.get', {
        'IBLOCK_TYPE_ID': 'lists',
        'IBLOCK_ID': '29',
    })
    fields_info = dict()
    for key in fields_raw.keys():
        fields_info[fields_raw[key]['NAME']] = fields_raw[key]

    report_data = [list(fields_info.keys())[:-2] + [list(fields_info.keys())[-1]]]
    year_code = list(filter(lambda x: x[1] == str(datetime.now().year), fields_info['Год']['DISPLAY_VALUES_FORM'].items()))[0][0]
    elements = fast_bitrix.get_all('lists.element.get', {
        'IBLOCK_TYPE_ID': 'lists',
        'IBLOCK_ID': '29',
        'FILTER': {
            fields_info['Год']['FIELD_ID']: year_code,
        }
    })
    responsible_field_id = fields_info['Ответственный']['FIELD_ID']
    users_info = fast_bitrix.get_all('user.get', {
        'FILTER': {
            'ID': list(map(lambda x: get_element_value('Ответственный', fields_info, x), filter(lambda y: responsible_field_id in y, elements)))
        }
    })
    for month_id in fields_info['Месяц']['DISPLAY_VALUES_FORM']:
        month_field_id = fields_info['Месяц']['FIELD_ID']
        month_name = list(filter(lambda x: x[0] == month_id, fields_info['Месяц']['DISPLAY_VALUES_FORM'].items()))[0][1]
        month_elements = list(filter(lambda elem: month_field_id in elem and get_element_value('Месяц', fields_info, elem) == month_id, elements))
        if not month_elements:
            continue
        month_result = list()
        for elem in month_elements:
            responsible = list(filter(lambda x: x['ID'] == get_element_value('Ответственный', fields_info, elem), users_info))
            if not responsible:
                continue
            responsible = responsible[0]
            month_result.append([
                elem['NAME'],
                f"{responsible['LAST_NAME']} {responsible['NAME']}",
                get_element_value('Дата', fields_info, elem),
                month_name,
                datetime.now().year,
                get_element_value('Потребность ТР на месяц по КП', fields_info, elem),
                get_element_value('Всего на объекте', fields_info, elem),
                get_element_value('В работе', fields_info, elem),
                get_element_value('В простое', fields_info, elem),
                get_element_value('В ремонте', fields_info, elem),
                get_element_value('В перебазировке', fields_info, elem),
                get_element_value('В плане на списание', fields_info, elem),
                get_element_value('Сторонние ТР', fields_info, elem),
                get_element_value('ИТОГО работоспособных ТС на объекте', fields_info, elem),
                get_element_value('КТГ', fields_info, elem),
                get_element_value('КИП', fields_info, elem),
            ])
        try:
            report_data += (sorted(month_result, key=lambda item: datetime.strptime(item[2], '%d.%m.%Y',)))
        except ValueError:
            pass
        month_result = list()
    for row in report_data:
        sheet.append(row)

    # Изменение ширины ячеек
    for index, row in enumerate(sheet.columns):
        sheet.column_dimensions[get_column_letter(row[0].column)].width = 16

    # Автоперенос текста, стиль текста и выравнивание
    for row_index, row in enumerate(sheet.rows):
        for cell_index, cell in enumerate(row):
            if row_index < 1:
                cell.font = Font(name='Calibri', size=12, bold=True)
            cell.alignment = cell.alignment.copy(wrapText=True)
            if cell_index > 1:
                cell.alignment = Alignment(horizontal='center')


    report_name = f'УС_РАС_{datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}.xlsx'
    with NamedTemporaryFile() as temp:
        book.save(temp.name)
        temp.seek(0)
        report_file = temp.read()
    report_file_base64 = str(base64.b64encode(report_file))[2:]

    # Загрузка отчета в Битрикс
    for user_id, user_email in USER_LIST.items():
        bitrix_folder_id = get_user_folder_id(fast_bitrix, user_id)
        upload_report = fast_bitrix.call('disk.folder.uploadfile', {
            'id': bitrix_folder_id,
            'data': {'NAME': report_name},
            'fileContent': report_file_base64
        })

        message_text = f'Отчет по УС:РАС сформирован.\nСсылка на отчет в Битрикс24: {upload_report["DETAIL_URL"]}'

        fast_bitrix.call('im.notify.system.add', {
            'USER_ID': user_id,
            'MESSAGE': f'Отчет по УС:РАС сформирован. {upload_report["DETAIL_URL"]}'}, raw=True)

        message = MIMEText(message_text)
        message['From'] = formataddr(('Робот Б24', 'robot_bitrix24@GSPROM.RU'))
        message['Subject'] = f'Отчет "Ремонт автотранспортных средств" {datetime.now().strftime("%d.%m.%Y")}'
        message['Message-ID'] = make_msgid()
        server = smtplib.SMTP('email.gsprom.ru:587')
        server.ehlo()
        server.starttls()
        server.ehlo
        server.login('robot_bitrix24@GSPROM.RU', 'Wv6zh8yw')
        message['To'] = user_email
        server.sendmail(message['From'], message['To'], message.as_string())
        server.quit()


if __name__ == '__main__':
    excel_list_28()
