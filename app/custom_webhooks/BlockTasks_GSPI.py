from datetime import timedelta, date, datetime
from tempfile import NamedTemporaryFile
import base64

from fastapi import Request
import openpyxl
from openpyxl.utils import get_column_letter

from custom_webhooks.ExcelList_28 import get_user_folder_id



async def create_report(fast_bitrix, params: Request.query_params):
    start_week = date.today()
    while start_week.isoweekday() != 1:
        start_week = start_week - timedelta(days=1)
    end_week = start_week + timedelta(days=6)
    current_week_completed_tasks = await fast_bitrix.get_all(
        'tasks.task.list',
        {
            'filter': {
                'GROUP_ID': '417',
                'STATUS': '5',
                '>DEADLINE': start_week.isoformat(),
                '<DEADLINE': end_week.isoformat(),
            }
        }
    )
    report_titles = [
        ['ГСП-Информсервис'],
        ['Блок СиРИС-Автоматизация (Романов Е.А.)'],
    ]
    book = openpyxl.Workbook()
    sheet = book.active
    for title in report_titles:
        sheet.append(title)
    sheet.append([f"Выполнено {start_week.strftime('%d.%m.%Y')}-{end_week.strftime('%d.%m.%Y')}"])
    sheet.append([''])
    users = set(map(lambda task: task['responsibleId'], current_week_completed_tasks))
    departments_users = dict()
    departments_name = dict()
    for user in users:
        user_info = await fast_bitrix.get_all('user.get', {'ID': user})
        user_info = user_info[0]
        departments_users.setdefault(user_info['UF_DEPARTMENT'][0], []).append(user)
        if user_info['UF_DEPARTMENT'][0] not in departments_name:
            department_info = await fast_bitrix.get_all('department.get', {'ID': user_info['UF_DEPARTMENT'][0]})
            department_info = department_info[0]
            departments_name[department_info['ID']] = department_info['NAME']
    for department in departments_users:
        department_tasks = list(filter(lambda task: task['responsibleId'] in departments_users[department], current_week_completed_tasks))
        sheet.append([departments_name[str(department)]])
        for task in department_tasks:
            sheet.append([
                '', task['title']
            ])
        sheet.append([''])

    for index, column_cells in enumerate(sheet.columns):
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = 50
    sheet.title = 'Выполнено'

    sheet = book.create_sheet('Запланировано')
    start_week = end_week
    end_week = start_week + timedelta(days=6)

    next_week_completed_tasks = await fast_bitrix.get_all(
        'tasks.task.list',
        {
            'filter': {
                'GROUP_ID': '417',
                '>DEADLINE': start_week.isoformat(),
                '<DEADLINE': end_week.isoformat(),
            }
        }
    )
    for title in report_titles:
        sheet.append(title)
    sheet.append(['Запланировано', f"{start_week.strftime('%d.%m.%Y')}-{end_week.strftime('%d.%m.%Y')}"])
    sheet.append([''])
    users = set(map(lambda task: task['responsibleId'], next_week_completed_tasks))
    departments_users = dict()
    departments_name = dict()
    for user in users:
        user_info = await fast_bitrix.get_all('user.get', {'ID': user})
        user_info = user_info[0]
        departments_users.setdefault(user_info['UF_DEPARTMENT'][0], []).append(user)
        if user_info['UF_DEPARTMENT'][0] not in departments_name:
            department_info = await fast_bitrix.get_all('department.get', {'ID': user_info['UF_DEPARTMENT'][0]})
            department_info = department_info[0]
            departments_name[department_info['ID']] = department_info['NAME']
    for department in departments_users:
        department_tasks = list(filter(lambda task: task['responsibleId'] in departments_users[department], next_week_completed_tasks))
        sheet.append([departments_name[str(department)]])
        for task in department_tasks:
            sheet.append([
                '', task['title'],
            ])
        sheet.append([''])
    for index, column_cells in enumerate(sheet.columns):
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = 50

    report_name = f'Автоматизация_{start_week}-{end_week}_{datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}.xlsx'
    with NamedTemporaryFile() as temp:
        book.save(temp.name)
        temp.seek(0)
        stream = temp.read()

    bitrix_folder_id = await get_user_folder_id(fast_bitrix, params['user_id'][5:], 'План_работ')
    report_file_base64 = str(base64.b64encode(stream))[2:]
    upload_report = await fast_bitrix.call('disk.folder.uploadfile', {
        'id': bitrix_folder_id,
        'data': {'NAME': report_name},
        'fileContent': report_file_base64
    })

    await fast_bitrix.call('im.notify.system.add', {
        'USER_ID': params['user_id'][5:],
        'MESSAGE': f'Отчет по плану работ: {upload_report["DETAIL_URL"]}'}, raw=True)
