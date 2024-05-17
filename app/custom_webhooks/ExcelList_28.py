from datetime import datetime
import base64
import os

from fastapi import Request
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment


def get_element_value(field_name: str, fields_info: dict, element_info: dict) -> str:
    try:
        value = list(element_info[fields_info[field_name]['FIELD_ID']].values())[0]
        if '|RUB' in value:
            value = value.replace('|RUB', '')
        return value
    except KeyError:
        return '0'
    except AttributeError:
        return element_info[fields_info[field_name]['FIELD_ID']]


def get_indicator_color(main_value: int, other_value: int) -> str:
    try:
        percent_value = int(other_value * 100 / main_value)
    except ZeroDivisionError:
        return 'ff0000'
    if percent_value < 50:
        return 'ff0000'
    elif percent_value < 80:
        return 'ffbf00'
    else:
        return '00b04f'


async def get_user_folder_id(fast_bitrix, user_id: str, folder_name: str = 'РАС') -> str:
    storage_info = await fast_bitrix.get_all('disk.storage.getlist', {
        'filter': {
            'ENTITY_TYPE': 'user',
            'ENTITY_ID': user_id,
        }
    })

    storage_id = storage_info[0]['ID']
    storage_folders = await fast_bitrix.get_all('disk.storage.getchildren', {
        'id': storage_id
    })
    report_folder = list(filter(lambda x: x['NAME'] == folder_name, storage_folders))
    if report_folder:
        return report_folder[0]['ID']

    new_folder = await fast_bitrix.call('disk.storage.addfolder', {
        'id': storage_id,
        'data': {
            'NAME': folder_name
        }
    })
    return new_folder['ID']


async def excel_list_28(fast_bitrix, params: Request.query_params):
    """
    Формирование отчета в формате .xlsx из данных списка и загрузка файла на диск.
    Пользователю, запустившему процесс, приходит ссылка на отчет в уведомлениях.

    :param fast_bitrix: Асинхронный экземпляр класса Bitrix библиотеки fast_bitrix24
    :param params: Параметры запроса:
                    list_id: ID списка
                    year: Год отчета
    """

    fields_raw = await fast_bitrix.get_all('lists.field.get', {
        'IBLOCK_TYPE_ID': 'lists',
        'IBLOCK_ID': params['list_id'],
    })

    fields_info = dict()
    for key in fields_raw.keys():
        fields_info[fields_raw[key]['NAME']] = fields_raw[key]

    year_code = list(filter(lambda x: x[1] == params['year'], fields_info['Год']['DISPLAY_VALUES_FORM'].items()))[0][0]
    elements = await fast_bitrix.get_all('lists.element.get', {
        'IBLOCK_TYPE_ID': 'lists',
        'IBLOCK_ID': params['list_id'],
        'FILTER': {
            fields_info['Год']['FIELD_ID']: year_code,
        }
    })
    responsible_field_id = fields_info['Ответственный']['FIELD_ID']
    users_info = await fast_bitrix.get_all('user.get', {
        'FILTER': {
            'ID': list(map(lambda x: get_element_value('Ответственный', fields_info, x), filter(lambda y: responsible_field_id in y, elements)))
        }
    })

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Ремонт автотранспортных средств'

    column_names = [
        [
            '' for _ in range(28)
        ],
        [
        '№ п/п', '№ п/п', 'Проект', 'Ответственный', 'Месяц', 'Год', 'План ремонта ТС, шт', 'Фактический план ремонта',
        'Дефектовка ТС, шт', 'Дефектовка, %', '',
        'Заявка на ремонт ТС, шт', 'Заявка на ремонт, %', '',
        'Закупка ЗЧ для ТС, шт', 'Закупка ЗЧ для ТС, %', '',
        'Оплата ЗЧ для ТС, шт', 'Оплата ЗЧ для ТС, %', '',
        'Потребность персонала РММ, чел.', 'Факт персонала РММ, чел.',
        'Поставка ЗЧ для ТС, шт', 'Поставка ЗЧ для ТС, %', '',
        'Ремонт ТС, шт', 'Ремонт ТС факт, %', '',
        'План ремонта ТС, шт', 'Ремонт ТС, шт', 'Ремонт ТС факт (долг), %',
        'план', 'факт',
        'план', 'факт',
        'факт', 'Остаток на р/с'
        ]
    ]
    column_names[0] += [
        'Перенос с предыдущего периода - Долг',
        '',
        '',
        'ОБС с учетом кредиторской задолженности',
        '',
        'ОперШтаб',
        '',
        'Мониторинговый счет',
    ]
    for names in column_names:
        sheet.append(names)

    year_element = list(filter(lambda x: x['NAME'] == 'Итого за год', elements))[0]
    sheet.append(
        [
            '',
            '',
            '',
            year_element['NAME'],
            'Все месяцы',
            params['year'],
            get_element_value('План ремонта ТС, шт', fields_info, year_element),
            '',
            get_element_value('Дефектовка ТС, шт', fields_info, year_element),
            '',
            '',
            get_element_value('Заявка на ремонт ТС, шт', fields_info, year_element),
            '',
            '',
            get_element_value('Закупка ЗЧ для ТС, шт', fields_info, year_element),
            '',
            '',
            get_element_value('Оплата ЗЧ для ТС, шт', fields_info, year_element),
            '',
            '',
            get_element_value('Потребность персонала РММ, чел.', fields_info, year_element),
            get_element_value('Факт персонала РММ, чел.', fields_info, year_element),
            get_element_value('Поставка ЗЧ для ТС, шт', fields_info, year_element),
            '',
            '',
            get_element_value('Ремонт ТС, шт', fields_info, year_element),
        ]
    )
    report_data = list()
    month_summary_rows = list()
    month_summary_rows_corrector = int()
    year_counter = 1
    for month_id in fields_info['Месяц']['DISPLAY_VALUES_FORM']:
        month_counter = 1
        month_field_id = fields_info['Месяц']['FIELD_ID']
        month_name = list(filter(lambda x: x[0] == month_id, fields_info['Месяц']['DISPLAY_VALUES_FORM'].items()))[0][1]
        month_elements = list(filter(lambda elem: month_field_id in elem and get_element_value('Месяц', fields_info, elem) == month_id, elements))
        if not month_elements:
            continue
        for elem_index, elem in enumerate(sorted(filter(lambda x: fields_info['Ответственный']['FIELD_ID'] in x, month_elements), key=lambda x: x['NAME'])):
            responsible = list(filter(lambda x: x['ID'] == get_element_value('Ответственный', fields_info, elem), users_info))[0]
            report_data.append(
                [
                    year_counter,
                    month_counter,
                    elem['NAME'],
                    f"{responsible['LAST_NAME']} {responsible['NAME']}",
                    month_name,
                    params['year'],
                    get_element_value('План ремонта ТС, шт', fields_info, elem),
                    get_element_value('Факт. план ремонта ТС, шт', fields_info, elem),
                    get_element_value('Дефектовка ТС, шт', fields_info, elem),
                    get_element_value('Дефектовка, %', fields_info, elem),
                    '',
                    get_element_value('Заявка на ремонт ТС, шт', fields_info, elem),
                    get_element_value('Заявка на ремонт, %', fields_info, elem),
                    '',
                    get_element_value('Закупка ЗЧ для ТС, шт', fields_info, elem),
                    get_element_value('Закупка ЗЧ для ТС, %', fields_info, elem),
                    '',
                    get_element_value('Оплата ЗЧ для ТС, шт', fields_info, elem),
                    get_element_value('Оплата ЗЧ для ТС, %', fields_info, elem),
                    '',
                    get_element_value('Потребность персонала РММ, чел.', fields_info, elem),
                    get_element_value('Факт персонала РММ, чел.', fields_info, elem),
                    get_element_value('Поставка ЗЧ для ТС, шт', fields_info, elem),
                    get_element_value('Поставка ЗЧ для ТС, %', fields_info, elem),
                    '',
                    get_element_value('Ремонт ТС, шт', fields_info, elem),
                    get_element_value('Ремонт ТС факт, %', fields_info, elem),
                    '',
                    get_element_value('План ремонта, шт. ДОЛГ', fields_info, elem),
                    get_element_value('Ремонт ТС, шт. ДОЛГ', fields_info, elem),
                    get_element_value('Ремонт ТС факт, % ДОЛГ', fields_info, elem),
                    get_element_value('Новое ОБС с учетом КЗ, План', fields_info, elem),
                    get_element_value('Новое ОБС с учетом КЗ, Факт', fields_info, elem),
                    get_element_value('ОперШтаб, План', fields_info, elem),
                    get_element_value('ОперШтаб, Факт', fields_info, elem),
                    get_element_value('Мониторинговый счет, Факт', fields_info, elem),
                    get_element_value('Мониторинговый счет, Остаток на р/сч', fields_info, elem),
                ]
            )
            year_counter += 1
            month_counter += 1
        try:
            month_summary_elem = list(filter(lambda x: x['NAME'] == 'Итого за месяц', month_elements))[0]
        except:
            continue
        report_data.append(
            [
                '',
                '',
                '',
                'Итого',
                month_name,
                params['year'],
                get_element_value('План ремонта ТС, шт', fields_info, month_summary_elem),
                get_element_value('Факт. план ремонта ТС, шт', fields_info, month_summary_elem),
                get_element_value('Дефектовка ТС, шт', fields_info, month_summary_elem),
                '',
                '',
                get_element_value('Заявка на ремонт ТС, шт', fields_info, month_summary_elem),
                '',
                '',
                get_element_value('Закупка ЗЧ для ТС, шт', fields_info, elem),
                '',
                '',
                get_element_value('Оплата ЗЧ для ТС, шт', fields_info, elem),
                '',
                '',
                get_element_value('Потребность персонала РММ, чел.', fields_info, elem),
                get_element_value('Факт персонала РММ, чел.', fields_info, elem),
                get_element_value('Поставка ЗЧ для ТС, шт', fields_info, elem),
                '',
                '',
                get_element_value('Ремонт ТС, шт', fields_info, elem),
                '',
                '',
            ]
        )
        month_summary_rows_corrector += 1
        month_summary_rows.append(year_counter + month_summary_rows_corrector)

    for row in report_data:
        sheet.append(row)

    indicator_indexes = {
        10: (6, 8),
        13: (6, 11),
        16: (6, 14),
        19: (6, 17),
        24: (6, 22),
        27: (6, 25),
    }

    # Изменение ширины ячеек
    for index, row in enumerate(sheet.columns):
        if index in list(indicator_indexes.keys()):
            sheet.column_dimensions[get_column_letter(row[0].column)].width = 2
        else:
            sheet.column_dimensions[get_column_letter(row[0].column)].width = 16

    # Автоперенос текста, стиль текста и выравнивание
    for row_index, row in enumerate(sheet.rows):
        for cell_index, cell in enumerate(row):
            if cell_index > 3 or row_index == 0:
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='left')

            if row_index < 2 or row_index - 1 in month_summary_rows:
                cell.font = Font(name='Calibri', size=12, bold=True)

            cell.alignment = cell.alignment.copy(wrapText=True)

    #Цвет ячеек
    for row_index, row in enumerate(sheet.rows):
        for cell_index, cell in enumerate(row):
            if row_index == 0:
                cell.fill = PatternFill(start_color='b4c6e7', end_color='b4c6e7', fill_type='solid')
            elif row_index == 1:
                cell.fill = PatternFill(start_color='c6e0b4', end_color='c6e0b4', fill_type='solid')
            elif 'Итого' in sheet.cell(row=row_index + 1, column=4).value:
                cell.fill = PatternFill(start_color='d9d9d9', end_color='d9d9d9', fill_type='solid')
            elif cell_index in list(indicator_indexes.keys()):
                try:
                    color = get_indicator_color(int(report_data[row_index - 1][indicator_indexes[cell_index][0]]), int(report_data[row_index - 1][indicator_indexes[cell_index][1]]))
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                except IndexError:
                    pass

    sheet.merge_cells('AC1:AE1')
    sheet.merge_cells('AF1:AG1')
    sheet.merge_cells('AH1:AI1')
    sheet.merge_cells('AJ1:AK1')

    sheet = workbook.create_sheet('Комплектация хабов тех.ресурсами')
    fields_raw = await fast_bitrix.get_all('lists.field.get', {
        'IBLOCK_TYPE_ID': 'lists',
        'IBLOCK_ID': '29',
    })
    fields_info = dict()
    for key in fields_raw.keys():
        fields_info[fields_raw[key]['NAME']] = fields_raw[key]

    report_data = [list(fields_info.keys())[:-1], ]
    year_code = list(filter(lambda x: x[1] == params['year'], fields_info['Год']['DISPLAY_VALUES_FORM'].items()))[0][0]
    elements = await fast_bitrix.get_all('lists.element.get', {
        'IBLOCK_TYPE_ID': 'lists',
        'IBLOCK_ID': '29',
        'FILTER': {
            fields_info['Год']['FIELD_ID']: year_code,
        }
    })
    responsible_field_id = fields_info['Ответственный']['FIELD_ID']
    users_info = await fast_bitrix.get_all('user.get', {
        'FILTER': {
            'ID': list(map(lambda x: get_element_value('Ответственный', fields_info, x), filter(lambda y: responsible_field_id in y, elements)))
        }
    })
    for month_id in fields_info['Месяц']['DISPLAY_VALUES_FORM']:
        month_field_id = fields_info['Месяц']['FIELD_ID']
        month_name = list(filter(lambda x: x[0] == month_id, fields_info['Месяц']['DISPLAY_VALUES_FORM'].items()))[0][1]
        month_elements = list(filter(lambda elem: month_field_id in elem and get_element_value('Месяц', fields_info, elem) == month_id, elements))
        if not month_elements:
            continue
        for elem in month_elements:
            responsible = list(filter(lambda x: x['ID'] == get_element_value('Ответственный', fields_info, elem), users_info))[0]
            report_data.append([
                elem['NAME'],
                f"{responsible['LAST_NAME']} {responsible['NAME']}",
                get_element_value('Дата', fields_info, elem),
                month_name,
                params['year'],
                get_element_value('Потребность ТР на месяц по КП', fields_info, elem),
                get_element_value('Всего на объекте', fields_info, elem),
                get_element_value('В работе', fields_info, elem),
                get_element_value('В простое', fields_info, elem),
                get_element_value('В ремонте', fields_info, elem),
                get_element_value('В перебазировке', fields_info, elem),
                get_element_value('В плане на списание', fields_info, elem),
                get_element_value('Сторонние ТР', fields_info, elem),
                get_element_value('ИТОГО работоспособных ТС на объекте', fields_info, elem),
            ])

    for row in report_data:
        sheet.append(row)

    # Изменение ширины ячеек
    for index, row in enumerate(sheet.columns):
            sheet.column_dimensions[get_column_letter(row[0].column)].width = 16

    # Автоперенос текста, стиль текста и выравнивание
    for row_index, row in enumerate(sheet.rows):
        for cell_index, cell in enumerate(row):
            if row_index < 1:
                cell.font = Font(name='Calibri', size=12, bold=True)
            cell.alignment = cell.alignment.copy(wrapText=True)


    report_name = f'УС_РАС_{datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}.xlsx'
    workbook.save(report_name)

    # Загрузка отчета в Битрикс
    bitrix_folder_id = await get_user_folder_id(fast_bitrix, params['user'][5:])
    with open(report_name, 'rb') as file:
        report_file = file.read()
    report_file_base64 = str(base64.b64encode(report_file))[2:]
    upload_report = await fast_bitrix.call('disk.folder.uploadfile', {
        'id': bitrix_folder_id,
        'data': {'NAME': report_name},
        'fileContent': report_file_base64
    })

    await fast_bitrix.call('im.notify.system.add', {
        'USER_ID': params['user'][5:],
        'MESSAGE': f'Отчет по УС:РАС сформирован. {upload_report["DETAIL_URL"]}'}, raw=True)
    os.remove(report_name)
