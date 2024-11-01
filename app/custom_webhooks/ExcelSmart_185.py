from datetime import datetime
import base64
import os
import re

from fastapi import Request
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from custom_webhooks.ExcelList_28 import get_element_value, get_user_folder_id

users_info_collection = list()


async def get_user_name(user_id, fast_bitrix):
    global users_info_collection
    user_info = list(filter(lambda x: x['ID'] == user_id, users_info_collection))
    if not user_info:
        user_info = await fast_bitrix.get_all('user.get', {'ID': user_id})
        users_info_collection += user_info
    try:
        return f"{user_info[0]['LAST_NAME']} {user_info[0]['NAME']} {user_info[0]['SECOND_NAME']}"
    except IndexError:
        return user_id


def get_enumeration_field_value(value, field_info):
    try:
        return list(filter(lambda x: str(x['ID']) == str(value), field_info['items']))[0]['VALUE']
    except IndexError:
        return value


async def excel_smart_185(fast_bitrix, params: Request.query_params):
    items = await fast_bitrix.get_all('crm.item.list', {
        'entityTypeId': '185',
        'select': [
            'title',
            'stageId',
            'ufCrm5_1712046836',
            'ufCrm5_1712047958',
            'ufCrm5_1712049480',
            'ufCrm5_1712049395',
            'ufCrm5_1712049436',
            'ufCrm5_1712648792',
            'ufCrm5_1712301306',
            'ufCrm5_1712049544',
            'ufCrm5_1712049569',
            'ufCrm5_1712049616',
            'ufCrm5_1712049659',
            'ufCrm5_1712049672',
            'ufCrm5_1712051323',
            'ufCrm5_1712049773',
            'ufCrm5_1712049833',
            'ufCrm5_1712049853',
            'ufCrm5_1712049891',
            'ufCrm5_1714650169',
            'ufCrm5_1712049593',
            'ufCrm5_1712049741',
            'ufCrm5_1716538019',
            'ufCrm5_1716538032',
            'ufCrm5_1716796215',
            'ufCrm5_1716537997',
            'ufCrm5_1721914677327',
            'ufCrm5_1721915566264',
            'ufCrm5_1721915580214',
            'ufCrm5_1721915591798',
            'ufCrm5_1721913228187',
            'ufCrm5_1721913267676',
            'ufCrm5_1721913304208',
            'ufCrm5_1721913336344',
            'ufCrm5_1721914677327',
            'ufCrm5_1721914704672',
            'ufCrm5_1721914738682',
            'ufCrm5_1721915610894',
            'ufCrm5_1721915760913',
        ]
    })

    users_info = list()

    fields_info = await fast_bitrix.get_all('crm.item.fields', {
        'entityTypeId': '185',
    })
    fields_info = fields_info['fields']

    status_dict = {
        'DT185_9:NEW': 'В работе',
        'DT185_9:SUCCESS': 'Завершено',
        'DT185_9:CLIENT': 'Утверждение УПАиК'
    }

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.append(
        [
            '',
            '',
            '',
            '',
            '',
            '',
            'Текущее состояние',
            '',
            '',
            '',
            '',
            '',
            'РЕЗУЛЬТАТ ПРЕДЫДУЩЕГО МЕСЯЦА, ПО КОТОРОМУ ФОРМИРУЕТСЯ СПРАВКА',
            '',
            '',
            '',
            '',
            'НАКАПЛИВАЕМЫЕ ИСТОРИЧЕСКИЕ ДАННЫЕ',
            '',
            '',
            '',
            '',
        ]
    )
    sheet.append(
        [
            'Номер',
            'Наименование показателя',
            "Стадия",
            'Направление',
            'Ответственный за показатель',
            'Блок',
            'Приоритет',
            'Ответственный в УПАиК',
            'Текущее значение',
            'Результат в %',
            'Дата среза',
            'Целевой результат по итогам встречи (или установленный отдельно ГД ГСП на год)',
            'Основные задачи на текущую и последующую недели',
            'Целевой показатель текущего месяца',
            'Результат исполнения прошлых периодов',
            '',
            'Дополнительные комментарии',
            'Результат исполнения ЯНВАРЯ',
            'Результат исполнения ФЕВРАЛЯ',
            'Результат исполнения МАРТА',
            'Результат исполнения 1-го квартала',
            'Светофор 1-го квартала',
            'Результат исполнения АПРЕЛЬ',
            'Результат исполнения МАЙ',
            'Результат исполнения ИЮНЬ',
            'Результат исполнения 2-го квартала',
            'Светофор 2-го квартала',
            'Результат исполнения ИЮЛЬ',
            'Результат исполнения АВГУСТ',
            'Результат исполнения СЕНТЯБРЬ',
            'Результат исполнения 3-го квартала',
            'Светофор 3-го квартала',
            'Результат исполнения ОКТЯБРЬ',
            'Результат исполнения НОЯБРЬ',
            'Результат исполнения ДЕКАБРЬ',
            'Результат исполнения 4-го квартала',
            'Светофор 4-го квартала'
        ]
    )

    for row in sorted(items, key=lambda item: float(re.search(r'(\d+[.]?\d+)|(\d+)', item['title']).group())):
        sheet.append(
            [
                row['title'],
                get_enumeration_field_value(row['ufCrm5_1712046836'], fields_info['ufCrm5_1712046836']),
                status_dict[row['stageId']],
                get_enumeration_field_value(row['ufCrm5_1712047958'], fields_info['ufCrm5_1712047958']),
                await get_user_name(row['ufCrm5_1712049480'], fast_bitrix),
                get_enumeration_field_value(row['ufCrm5_1712049395'], fields_info['ufCrm5_1712049395']),
                get_enumeration_field_value(row['ufCrm5_1712049436'], fields_info['ufCrm5_1712049436']),
                await get_user_name(row['ufCrm5_1712301306'], fast_bitrix),
                row['ufCrm5_1712049544'],
                row['ufCrm5_1712049569'],
                datetime.fromisoformat(row['ufCrm5_1712049593']).strftime('%d.%m.%Y') if row['ufCrm5_1712049593'] else '',
                row['ufCrm5_1712049616'],
                row['ufCrm5_1712049659'],
                row['ufCrm5_1712049672'],
                row['ufCrm5_1712051323'],
                get_enumeration_field_value(row['ufCrm5_1712049741'], fields_info['ufCrm5_1712049741']),
                row['ufCrm5_1712049773'],
                row['ufCrm5_1712049833'],
                row['ufCrm5_1712049853'],
                row['ufCrm5_1712049891'],
                row['ufCrm5_1721913228187'],
                row['ufCrm5_1721913267676'],
                row['ufCrm5_1714650169'],
                row['ufCrm5_1716538019'],
                row['ufCrm5_1716538032'],
                row['ufCrm5_1716796215'],
                row['ufCrm5_1716537997'],
                row['ufCrm5_1721914677327'],
                row['ufCrm5_1721915566264'],
                row['ufCrm5_1721915580214'],
                row['ufCrm5_1721915591798'],
            ]
        )

    cell_indicator_colors = {
        'зеленый': '00b04f',
        'желтый': 'ffbf00',
        'красный': 'ff0000'
    }

    row_cell_width = {
        0: 10,
        1: 50,
        2: 15,
        3: 25,
        4: 35,
        5: 25,
        6: 12,
        7: 35,
        8: 40,
        9: 15,
        10: 12,
        11: 50,
        12: 50,
        13: 50,
        14: 50,
        15: 5,
        16: 50,
        17: 50,
        18: 50,
        19: 50,
        20: 50,
        21: 50,
        22: 50,
        23: 50,
        24: 50,
        25: 50,
        26: 50,
        27: 50,
        28: 50,
        29: 50,
        30: 50,
        31: 50,
        32: 50,
    }

    # Изменение ширины ячеек
    for index, row in enumerate(sheet.columns):
        if index in row_cell_width:
            cell_width = row_cell_width[index]
        else:
            cell_width = 50
        sheet.column_dimensions[get_column_letter(row[0].column)].width = cell_width

    # Автоперенос текста, стиль текста и выравнивание
    for row_index, row in enumerate(sheet.rows):
        for cell_index, cell in enumerate(row):
            if row_index == 1 or (row_index == 0 and cell_index in [6, 7, 8,]):
                if row_index == 1 and cell_index in [12, 13, 17, 18, 19, 20, 21, 22, 23,]:
                    cell.fill = PatternFill(start_color='EAD7B8', end_color='EAD7B8', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='5EBDF7', end_color='5EBDF7', fill_type='solid')
                cell.font = Font(name='Calibri', size=12, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:

                # Цвет ячеек
                if row_index == 0 and cell_index in [12, 13, 17]:
                    cell.fill = PatternFill(start_color='EAD7B8', end_color='EAD7B8', fill_type='solid')
                elif cell_index == 15 and row_index > 1:
                    cell_color = cell_indicator_colors[cell.value.lower()] if cell.value in cell_indicator_colors else 'ff0000'
                    cell.fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type='solid')
                    cell.value = ''

                if cell_index in [0, 2, 3, 4, 5, 6, 7, 9, 10] or row_index == 0 and cell_index == 17:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

            cell.alignment = cell.alignment.copy(wrapText=True)
            cell.font = Font(name='Arial Narrow', size=12)
            cell.border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin')
                                 )
    sheet.merge_cells('G1:I1')
    sheet.merge_cells('M1:N1')
    sheet.merge_cells('R1:X1')

    report_name = f'Показатели_функционального_блока{datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}.xlsx'
    workbook.save(report_name)

    bitrix_folder_id = await get_user_folder_id(fast_bitrix, params['user'][5:], 'ПФБ')
    with open(report_name, 'rb') as file:
        report_file = file.read()
    report_file_base64 = str(base64.b64encode(report_file))[2:]
    upload_report = await fast_bitrix.call('disk.folder.uploadfile', {
        'id': bitrix_folder_id,
        'data': {'NAME': report_name},
        'fileContent': report_file_base64
    })

    await fast_bitrix.call('im.notify.system.add', {
        'USER_ID': params['user'][5:],
        'MESSAGE': f'Отчет по показателям функционального блока сформирован. {upload_report["DETAIL_URL"]}'}, raw=True)
    os.remove(report_name)

