from datetime import datetime, timedelta, date
import re
import base64
import locale

from fastapi import Request
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from tempfile import NamedTemporaryFile
import smtplib
from email.mime.text import MIMEText
from email.utils import make_msgid
from email.utils import formataddr

from custom_webhooks.ExcelList_28 import get_user_folder_id


locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')


async def send_notification(fast_bitrix, params: Request.query_params, notification_type: str = 'email'):
    start_week = date.today()
    while start_week.isoweekday() != 1:
        start_week = start_week - timedelta(days=1)
    end_week = start_week + timedelta(days=6)
    tasks = await fast_bitrix.get_all('tasks.task.list', {
        'filter': {
            'GROUP_ID': '402',
            '>END_DATE_PLAN': (start_week - timedelta(days=1)).isoformat(),
            '<END_DATE_PLAN': (end_week + timedelta(days=1)).isoformat(),
        }
    })
    users = set(map(lambda task: task['responsibleId'], tasks))
    for user in users:
        user_tasks = list(filter(lambda task: task['responsibleId'] == user and not task['closedDate'], tasks))
        if user_tasks:
            message_text = 'Данные задачи по плану работ не завершены:\n'
            for index, user_task in enumerate(user_tasks, 1):
                message_text += f'{index}. {user_task["title"]} https://bitrix.gsprom.ru/workgroups/group/402/tasks/task/view/{user_task["id"]}/\n'
            if notification_type == 'bitrix':
                await fast_bitrix.call('im.notify.system.add', {
                    'USER_ID': user,
                    'MESSAGE': message_text}, raw=True)
            elif notification_type == 'email':
                user_info = await fast_bitrix.get_all('user.get', {'ID': user})
                message = MIMEText(message_text)
                message['From'] = formataddr(('План мероприятий', '1c_b24_task_plan@gsprom.ru'))
                message['Subject'] = f'Контроль выполнения плана мероприятий за {start_week.strftime("%d.%m")} - {end_week.strftime("%d.%m")}'
                message['Message-ID'] = make_msgid()
                server = smtplib.SMTP('email.gsprom.ru:587')
                server.ehlo()
                server.starttls()
                server.ehlo
                server.login('1c_b24_task_plan@gsprom.ru', 'bk=/PN1F3Q%H')
                message['To'] = user_info[0]['EMAIL']
                server.sendmail(message['From'], message['To'], message.as_string())
                server.quit()

    await fast_bitrix.call('im.notify.system.add', {
        'USER_ID': params['user_id'][5:],
        'MESSAGE': 'Уведомления отправлены'}, raw=True)


async def create_report(fast_bitrix, params: Request.query_params):
    datetime_format = '%d.%m.%Y'
    tasks = await fast_bitrix.get_all(
        'tasks.task.list',
        {
            'filter': {
                'GROUP_ID': '402',
                '>END_DATE_PLAN': (datetime.strptime(params['date_filter_start'], '%d.%m.%Y') - timedelta(days=1)).isoformat(),
                '<END_DATE_PLAN': (datetime.strptime(params['date_filter_end'], '%d.%m.%Y') + timedelta(days=1)).isoformat(),
            }
        }
    )
    if params['report_type'] == 'users':
        group_users = await fast_bitrix.get_all(
            'sonet_group.user.get',
            {
                'ID': '402',
            }
        )
        non_task_users = set(map(lambda users: users['USER_ID'], group_users)) - set(map(lambda task: task['responsibleId'], tasks))
        users_info = await fast_bitrix.get_all(
            'user.get',
            {
                'filter': {
                    'ID': list(non_task_users)
                }
            }
        )
        message = (f"Пользователи, которые не создали задачи плана работ на даты {params['date_filter_start']} - {params['date_filter_end']}:\n")
        for index, user in enumerate(users_info, 1):
            message += f"{user['LAST_NAME']} {user['NAME'][0]}\n"
        await fast_bitrix.call('im.notify.system.add', {
            'USER_ID': params['user_id'][5:],
            'MESSAGE': message
        }
                               )

    users_info = await fast_bitrix.get_all(
        'user.get',
        {
            'filter': {
                'ID': list(map(lambda task: task['responsibleId'], tasks))
            }
        }
    )

    start_datetime = datetime.strptime(params['date_filter_start'], '%d.%m.%Y')
    report_data = [
        [],
        [
            '',
            f"ЗАДАЧИ {start_datetime.strftime('%B %Y').upper()} г.",
        ],
        [
            '',
            'Организационная единица',
            'ГСП ИнформСервис',
        ],
        [
            '',
            'Исполнитель',
            'Чугуевский И.О.',
        ],
        [
            '',
            'Дата',
            datetime.now().strftime('%d.%m.%Y')
        ],
        [],
        [],
        [
            '№',
            'Функциональный блок',
            'Управление',
            'Отдел',
            'Тип задачи',
            'Задача',
            'Зачем',
            'Уровень контроля',
            'Ответственный',
            'Риски',
            'Необходимая поддержка',
            'Дата начала (план)',
            'Дата окончания (план)',
            'Время исполнения (план)',
            'Дата начала (факт)',
            'Дата окончания (факт)',
            'Время исполнения (факт)',
            'Статус на дату',
            'Исполнение срока на дату',
            'Комментарий'
        ]
    ]
    task_status = {
        '3': 'В работе',
        '5': 'Выполнено'
    }
    task_counter = 1
    success_rows = []
    in_work_rows = []
    for user in sorted(users_info, key=lambda data: data['LAST_NAME']):
        user_name = f"{user['LAST_NAME']} {user['NAME'][0]}."
        if user['SECOND_NAME']:
            user_name += user['SECOND_NAME'][0] + '.'
        user_tasks = filter(lambda task: task['responsibleId'] == user['ID'], tasks)
        user_department = await fast_bitrix.get_all(
            'department.get',
            {
                'ID': user['UF_DEPARTMENT'][0]
            }
        )
        try:
            user_parent_department = await fast_bitrix.get_all(
                'department.get',
                {
                    'ID': user_department[0]['PARENT'],
                }
            )
            user_parent_department_name = user_parent_department[0]['NAME']
        except:
            user_parent_department_name = ''
        try:
            user_head_department = await fast_bitrix.get_all(
                'department.get',
                {
                    'ID': user_parent_department[0]['PARENT'],
                }
            )
            user_head_department_name = user_head_department[0]['NAME']
        except:
            user_head_department_name = ''
        for task in sorted(user_tasks, key=lambda task: datetime.fromisoformat(task['createdDate'])):
            task_description = task['title']
            task_target = re.search(r'Зачем:\W(.*)', task['description'])
            print(task['id'], task_target)
            task_control_level = re.search(r'Уровень контроля:\W(.*)', task['description'])
            task_risk = re.search(r'Риски:\W(.*)', task['description'])
            task_support = re.search(r'Необходимая поддержка:\W(.*)', task['description'])
            plan_date_start = datetime.fromisoformat(task['startDatePlan']) if task['startDatePlan'] else None
            plan_date_end = datetime.fromisoformat(task['endDatePlan']) if task['endDatePlan'] else None
            fact_date_start = datetime.fromisoformat(task['dateStart']) if task['dateStart'] else None
            fact_date_end = datetime.fromisoformat(task['closedDate']) if task['closedDate'] else None
            if not fact_date_start and fact_date_end:
                fact_date_start = plan_date_start
            task_commentary = await fast_bitrix.call('tasks.task.result.list', {
                'taskId': task['id']
            })

            report_data.append(
                [
                    task_counter,
                    user_head_department_name,
                    user_parent_department_name,
                    user_department[0]['NAME'],
                    'Плановые',
                    task_description,
                    task_target.group(1).replace('--TEXT--', '') if task_target else None,
                    task_control_level.group(1).replace('--TEXT--', '') if task_control_level else None,
                    user_name,
                    task_risk.group(1).replace('--TEXT--', '') if task_risk else None,
                    task_support.group(1).replace('--TEXT--', '') if task_support else None,
                    plan_date_start.strftime(datetime_format) if plan_date_start else None,
                    plan_date_end.strftime(datetime_format) if plan_date_end else None,
                    (plan_date_end - plan_date_start).days if all([plan_date_end, plan_date_start]) else None,
                    fact_date_start.strftime(datetime_format) if fact_date_start else None,
                    fact_date_end.strftime(datetime_format) if fact_date_end else None,
                    (fact_date_end - fact_date_start).days + 1 if all([fact_date_end, fact_date_start]) else None,
                    task_status[task['status']] if task['status'] in task_status else 'Не начато',
                    'В срок' if all([plan_date_end, fact_date_end]) and plan_date_end.strftime(datetime_format) <= fact_date_end.strftime(datetime_format) else 'Не в срок',
                    task_commentary['text'] if task_commentary else None,

                ]
            )
            if task['status'] == '5':
                success_rows.append(task_counter)
            elif task['status'] == '3':
                in_work_rows.append(task_counter)
            task_counter += 1

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in report_data:
        sheet.append(row)

    custom_format = (
        (7, 'center'),      # №
        (40, 'center'),     # Функциональный блок
        (50, 'center'),     # Управление
        (40, 'center'),     # Отдел
        (20, 'center'),     # Тип задачи
        (60, 'left'),       # Задача
        (60, 'left'),       # Зачем
        (20, 'center'),     # Уровень контроля
        (30, 'center'),     # Ответственный
        (40, 'left'),       # Риски
        (30, 'left'),       # Необходимая поддержка
        (20, 'center'),     # Дата начала (план)
        (20, 'center'),     # Дата окончания (план)
        (20, 'center'),     # Время исполнения (план)
        (20, 'center'),     # Дата начала (факт)
        (20, 'center'),     # Дата окончания (факт)
        (20, 'center'),     # Время исполнения (факт)
        (20, 'center'),     # Статус на дату
        (20, 'center'),     # Исполнения срока на дату
        (50, 'left'),       # Комментарий
    )

    # Автоперенос текста, стиль текста и выравнивание
    for row_index, row in enumerate(sheet.rows):
        for cell_index, cell in enumerate(row):
            if row_index in (1, 2, 3, 4):
                if cell_index == 1:
                    cell.alignment = Alignment(horizontal='center', wrapText=True, vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='right', wrapText=True, vertical='center')
            elif row_index == 7:
                cell.alignment = Alignment(horizontal='center', wrapText=True, vertical='center')
            elif row_index > 7:
                cell.alignment = Alignment(horizontal=custom_format[cell_index][1], wrapText=True, vertical='center')

            # Границы ячеек
            if row_index <= 7:
                cell.border = Border(
                    left=Side('thin', color='FFFFFF'),
                    right=Side('thin', color='FFFFFF'),
                    top=Side('thin', color='FFFFFF'),
                    bottom=Side('thin', color='FFFFFF'),
                )
            else:
                cell.border = Border(
                    left=Side('thin', color='000000'),
                    right=Side('thin', color='000000'),
                    top=Side('thin', color='000000'),
                    bottom=Side('thin', color='000000'),
                )

    # Цвет ячеек
    for row_index, row in enumerate(sheet.rows):
        for cell_index, cell in enumerate(row):
            if row_index == 7:
                cell.fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
                cell.font = Font(name='Arial Narrow', color='FFFFFF', bold=True)
            elif row_index - 7 in success_rows:
                cell.fill = PatternFill(start_color='b0dc9c', end_color='b0dc9c', fill_type='solid')
            elif row_index - 7 in in_work_rows:
                cell.fill = PatternFill(start_color='f8b484', end_color='f8b484', fill_type='solid')
            elif row_index in (1, 2, 3, 4):
                if cell_index == 1:
                    cell.font = Font(name='Arial Narrow', bold=True, size=14)
                else:
                    cell.font = Font(name='Arial Narrow', size=14)

    # Изменение ширины ячеек
    for column_index, row in enumerate(sheet.columns):
        if column_index in (0, 5, 6):
            sheet.row_dimensions[column_index + 1].height = 0
        elif column_index in (1, 2, 3, 4):
            sheet.row_dimensions[column_index + 1].height = 25
        else:
            sheet.row_dimensions[column_index + 1].height = 50
        sheet.column_dimensions[get_column_letter(row[0].column)].width = custom_format[column_index][0]

    # Фикс ряда и фильтры
    sheet.freeze_panes = sheet['A9']
    sheet.auto_filter.ref = 'A8:T8'

    report_name = f'Отчет_по_плану_работ_{datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}.xlsx'
    with NamedTemporaryFile() as temp:
        workbook.save(temp.name)
        temp.seek(0)
        stream = temp.read()

    bitrix_folder_id = await get_user_folder_id(fast_bitrix, params['user_id'][5:], 'План_работ')
    report_file_base64 = str(base64.b64encode(stream))[2:]
    upload_report = await fast_bitrix.call('disk.folder.uploadfile', {
        'id': bitrix_folder_id,
        'data': {'NAME': report_name},
        'fileContent': report_file_base64
    })

    await fast_bitrix.call('im.notify.system.add', {
        'USER_ID': params['user_id'][5:],
        'MESSAGE': f'Отчет по плану работ: {upload_report["DETAIL_URL"]}'}, raw=True)


async def create_non_completed_tasks_report(fast_bitrix, params: Request.query_params):
    start_week = date.today()
    while start_week.isoweekday() != 1:
        start_week = start_week - timedelta(days=1)
    end_week = start_week + timedelta(days=6)
    tasks = await fast_bitrix.get_all('tasks.task.list', {
        'filter': {
            'GROUP_ID': '402',
            '>END_DATE_PLAN': (start_week - timedelta(days=1)).isoformat(),
            '<END_DATE_PLAN': (end_week + timedelta(days=1)).isoformat(),
        }
    })
    users = set(map(lambda task: task['responsibleId'], tasks))
    message_text = 'Те, кто еще не отчитался, а срок уже завтра:\n'
    for user in users:
        user_tasks = list(filter(lambda task: task['responsibleId'] == user and not task['closedDate'], tasks))
        if user_tasks:
            message_text += f"{user_tasks[0]['responsible']['name']}:\n"
            for index, user_task in enumerate(user_tasks, 1):
                message_text += f'{index}. {user_task["title"]} https://bitrix.gsprom.ru/workgroups/group/402/tasks/task/view/{user_task["id"]}/\n'
            message_text += '\n'
    if message_text == 'Те, кто еще не отчитался, а срок уже завтра:\n':
        message_text = 'Все задачи за неделю завершены'
    await fast_bitrix.call('im.notify.system.add', {
        'USER_ID': params['user_id'][5:],
        'MESSAGE': message_text}, raw=True)


