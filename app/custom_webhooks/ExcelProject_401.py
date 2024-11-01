import re
from tempfile import NamedTemporaryFile
from datetime import datetime, timedelta
import base64

from fastapi import Request
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from custom_webhooks.ExcelList_28 import get_user_folder_id
from custom_webhooks.ExcelSmart_185 import get_user_name




async def create_key_tasks_report(fast_bitrix, params: Request.query_params):
    tasks = await fast_bitrix.get_all(
        'tasks.task.list',
        {
            'filter': {
                '>CREATED_DATE': (datetime.strptime(params['date_filter_start'], '%d.%m.%Y') - timedelta(days=1)).isoformat(),
                '<CREATED_DATE': (datetime.strptime(params['date_filter_end'], '%d.%m.%Y') + timedelta(days=1)).isoformat(),
                'GROUP_ID': '401',
            }
        }
    )
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append([
        'п/п',
        'Тема (область)',
        'Формулировка задачи',
        'ФИО исполнителя',
        'Дата ожидаемого исполнения',
        'Ожидаемый результат',
        'Отчет (ход исполнения поручения)',
    ])
    for index, task in enumerate(tasks, 1):
        user_name = await get_user_name(task['responsible'], fast_bitrix)
        task_description = re.search(r'(Формулировка задачи:\W.*)\W', task['description'])
        if task_description:
            task_description = task_description.group().replace('Формулировка задачи:\n', '')
        else:
            task_description = ''
        task_result = re.search(r'(Ожидаемый результат:\W.*)', task['description'])
        if task_description:
            task_result = task_result.group().replace('Ожидаемый результат:\n', '')
        else:
            task_result = ''

        result_info = await fast_bitrix.call('tasks.task.result.list', {
            'taskId': task['id']
        })
        if result_info:
            result_info = result_info['text']
        else:
            result_info = ''
        sheet.append([
            index,
            task['title'],
            task_description.strip(),
            user_name,
            datetime.fromisoformat(task['deadline']).strftime('%d.%m.%Y'),
            task_result.strip(),
            result_info,
        ])

    cells_width = {
        0: 5,
        1: 25,
        2: 50,
        3: 25,
        4: 25,
        5: 50,
        6: 50,
    }

    for index, column_cells in enumerate(sheet.columns):
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = cells_width[index]

    for index, column_rows in enumerate(sheet.rows):
        for cell in column_rows:
            cell.alignment = cell.alignment.copy(wrapText=True)
            if index == 0:
                cell.alignment = Alignment(horizontal='center')

    with NamedTemporaryFile() as temp:
        workbook.save(temp.name)
        temp.seek(0)
        stream = temp.read()

    report_name = f'Отчет_по_ключевым_задачам_{datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}.xlsx'
    workbook.save(report_name)

    bitrix_folder_id = await get_user_folder_id(fast_bitrix, params['user_id'][5:], 'Ключевые_задачи')
    report_file_base64 = str(base64.b64encode(stream))[2:]
    upload_report = await fast_bitrix.call('disk.folder.uploadfile', {
        'id': bitrix_folder_id,
        'data': {'NAME': report_name},
        'fileContent': report_file_base64
    })

    await fast_bitrix.call('im.notify.system.add', {
        'USER_ID': params['user_id'][5:],
        'MESSAGE': f'Отчет по ключевым задачам: {upload_report["DETAIL_URL"]}'}, raw=True)


