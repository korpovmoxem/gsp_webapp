from datetime import datetime
import base64
import os

from fastapi import Request
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

from tools.tools import fast_bitrix_slow, fast_bitrix


def get_element_value(field_name: str, fields_info: dict, element_info: dict) -> str:
    try:
        value = list(element_info[fields_info[field_name]['FIELD_ID']].values())[0]
        if '|RUB' in value:
            value = value.replace('|RUB', '')
        return value
    except KeyError:
        return '0'
    except AttributeError:
        return element_info[fields_info[field_name]['FIELD_ID']]


def get_indicator_color(main_value: int, other_value: int) -> str:
    #try:
        #percent_value = int(int(other_value) * 100 / int(main_value))
    #except:
        #return 'ff0000'
    try:
        other_value = int(other_value)
    except:
        return 'ff0000'
    if other_value < 50:
        return 'ff0000'
    elif other_value < 80:
        return 'ffbf00'
    else:
        return '00b04f'

def get_element_row_value(elem, responsible, elem_type):
    row_values = [
            elem['NAME'],
            f"{responsible['LAST_NAME']} {responsible['NAME']}",
            month_name,
            2024,
            get_element_value('План ремонта ТС, шт', fields_info, elem),
            get_element_value('Перенос с предыдущего месяца', fields_info, elem),
            get_element_value('ИТОГО план ремонтов ТС на 1 число, шт', fields_info, elem),
            get_element_value('Дефектовка ТС, шт', fields_info, elem),
            get_element_value('Дефектовка, %', fields_info, elem),
            get_element_value('Заявка на ремонт ТС, шт', fields_info, elem),
            get_element_value('Заявка на ремонт, %', fields_info, elem),
            get_element_value('Закупка ЗЧ для ТС, шт', fields_info, elem),
            get_element_value('Закупка ЗЧ для ТС, %', fields_info, elem),
            get_element_value('Оплата ЗЧ для ТС, шт', fields_info, elem),
            get_element_value('Оплата ЗЧ для ТС, %', fields_info, elem),
            get_element_value('Поставка ЗЧ для ТС, шт', fields_info, elem),
            get_element_value('Поставка ЗЧ для ТС, %', fields_info, elem),
            get_element_value('Ремонт ТС, шт', fields_info, elem),
            get_element_value('Ремонт ТС факт, %', fields_info, elem),
            f"{round(float(get_element_value('ОБС с учетом КЗ, План', fields_info, elem))):_}".replace('_', ' '),
            f"{round(float(get_element_value('ОБС с учетом КЗ, Факт', fields_info, elem))):_}".replace('_', ' '),
            f"{round(float(get_element_value('ОперШтаб, План', fields_info, elem))):_}".replace('_', ' '),
            f"{round(float(get_element_value('ОперШтаб, Факт', fields_info, elem))):_}".replace('_', ' '),
            f"{round(float(get_element_value('Страховой запас, План', fields_info, elem))):_}".replace('_', ' '),
            f"{round(float(get_element_value('Страховой запас, Факт', fields_info, elem))):_}".replace('_', ' '),
            f"{round(float(get_element_value('Мониторинговый счет, Факт', fields_info, elem))):_}".replace('_', ' '),
            f"{round(float(get_element_value('Мониторинговый счет, Остаток на р/сч', fields_info, elem))):_}".replace('_', ' '),
            f"{round(float(get_element_value('Итого ФАКТ по хабу, руб.', fields_info, elem))):_}".replace('_', ' '),
        ]
    if elem['ID'] == '4338':
        print(elem)
        print(f"{round(float(get_element_value('ОБС с учетом КЗ, План', fields_info, elem))):_}".replace('_', ' '))
        print(f"{round(float(get_element_value('ОБС с учетом КЗ, Факт', fields_info, elem))):_}".replace('_', ' '))
        print(row_values)
    if elem_type == 'summary_month':
        row_values[1] = ''
        row_values[3] = ''
    return row_values



async def get_user_folder_id(fast_bitrix, user_id: str, folder_name: str = 'РАС') -> str:
    storage_info = await fast_bitrix.get_all('disk.storage.getlist', {
        'filter': {
            'ENTITY_TYPE': 'user',
            'ENTITY_ID': user_id,
        }
    })

    storage_id = storage_info[0]['ID']
    storage_folders = await fast_bitrix.get_all('disk.storage.getchildren', {
        'id': storage_id
    })
    report_folder = list(filter(lambda x: x['NAME'] == folder_name, storage_folders))
    if report_folder:
        return report_folder[0]['ID']

    new_folder = await fast_bitrix.call('disk.storage.addfolder', {
        'id': storage_id,
        'data': {
            'NAME': folder_name
        }
    })
    return new_folder['ID']


report_titles = [
    [
        'Проект', 'Ответственный', 'Месяц', 'Год', 'План ремонта ТС, шт', 'Перенос с предыдущего месяца', 'ИТОГО план ремонтов ТС на 1 число, шт',
        'Дефектовка ТС', '', 'Заявка на ремонт ТС', '', 'Закупка ЗЧ для ТС', '', 'Оплата ЗЧ для ТС', '', 'Поставка ЗЧ для ТС', '',
        'Ремонт ТС', '', 'ОБС с учетом КЗ, руб.', '', 'ОперШтаб, руб.', '', 'Страховой запас', '', 'Мониторинговый счет, Факт',
        'Мониторинговый счет, Остаток на р/сч', 'Итого ФАКТ по хабу, руб'
    ],
    [
        '', '', '', '', '', '', '', 'шт.', '%', 'шт.', '%', 'шт.', '%', 'шт.', '%', 'шт.', '%', 'шт.', '%',
        'План', 'Факт', 'План', 'Факт', 'План', 'Факт',
    ]
]
first_row_title = ['Свод по ремонтам технических ресурсов в разрезе Хаба'] + [''] * (len(report_titles[0]) - 3) + [datetime.now().strftime('%d.%m.%Y')]
report_titles.insert(0, first_row_title)

book = openpyxl.Workbook()
sheet = book.active

for row in report_titles:
    sheet.append(row)

sheet.merge_cells('A2:A3')      # Проект
sheet.merge_cells('B2:B3')      # Ответственный
sheet.merge_cells('C2:C3')      # Месяц
sheet.merge_cells('D2:D3')      # Год
sheet.merge_cells('E2:E3')      # План ремонта ТС, шт
sheet.merge_cells('F2:F3')      # Перенос с предыдущего месяца
sheet.merge_cells('G2:G3')      # ИТОГО план ремонтов ТС на 1 число, шт
sheet.merge_cells('A1:Z1')      # Свод по ремонтам технических ресурсов в разрезе Хаба
sheet.merge_cells('AA1:AB1')    # datetime.now()
sheet.merge_cells('H2:I2')      # Дефектовка ТС
sheet.merge_cells('J2:K2')      # Заявка на ремонт ТС
sheet.merge_cells('L2:M2')      # Закупка ЗЧ для ТС
sheet.merge_cells('N2:O2')      # Оплата ЗЧ для ТС
sheet.merge_cells('P2:Q2')      # Поставка ЗЧ для ТС
sheet.merge_cells('R2:S2')      # Ремонт ТС
sheet.merge_cells('T2:U2')      # ОБС с учетом КЗ, руб.
sheet.merge_cells('V2:W2')      # ОперШтаб, руб.
sheet.merge_cells('X2:Y2')      # Страховой запас
sheet.merge_cells('Z2:Z3')      # ОБС с учетом КЗ, руб.
sheet.merge_cells('AA2:AA3')    # ОперШтаб, руб.
sheet.merge_cells('AB2:AB3')    # Страховой запас

fields_raw = fast_bitrix_slow.get_all('lists.field.get', {
    'IBLOCK_TYPE_ID': 'lists',
    'IBLOCK_ID': '28',
})

fields_info = dict()
for key in fields_raw.keys():
    fields_info[fields_raw[key]['NAME']] = fields_raw[key]

year_code = list(filter(lambda x: x[1] == '2024', fields_info['Год']['DISPLAY_VALUES_FORM'].items()))[0][0]
elements = fast_bitrix_slow.get_all('lists.element.get', {
    'IBLOCK_TYPE_ID': 'lists',
    'IBLOCK_ID': '28',
    'FILTER': {
        fields_info['Год']['FIELD_ID']: year_code,
    }
})
responsible_field_id = fields_info['Ответственный']['FIELD_ID']
users_info = fast_bitrix_slow.get_all('user.get', {
    'FILTER': {
        'ID': list(map(lambda x: get_element_value('Ответственный', fields_info, x),
                       filter(lambda y: responsible_field_id in y, elements)))
    }
})

for month_id in fields_info['Месяц']['DISPLAY_VALUES_FORM']:
    month_field_id = fields_info['Месяц']['FIELD_ID']
    month_name = list(filter(lambda x: x[0] == month_id, fields_info['Месяц']['DISPLAY_VALUES_FORM'].items()))[0][1]
    month_elements = list(filter(lambda elem: month_field_id in elem and get_element_value('Месяц', fields_info, elem) == month_id, elements))
    if not month_elements:
        continue
    for elem_index, elem in enumerate(sorted(filter(lambda x: fields_info['Ответственный']['FIELD_ID'] in x, month_elements), key=lambda x: x['NAME'])):
        responsible = list(filter(lambda x: x['ID'] == get_element_value('Ответственный', fields_info, elem), users_info))[0]
        sheet.append(get_element_row_value(elem, responsible, 'row'))
    summary_month_element = list(filter(lambda element: get_element_value('Месяц', fields_info, element) == month_id and 'Итого за месяц' in element['NAME'], month_elements))
    if not summary_month_element:
        sheet.append([])
        continue
    sheet.append(get_element_row_value(summary_month_element[0], responsible, 'summary_month'))
    for col_index, _ in enumerate(next(sheet.rows), 1):
        cell = sheet.cell(len(list(sheet.rows)), col_index)
        cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        cell.font = Font(bold=True)


thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

for row_index, row in enumerate(sheet.rows, 1):
    for cell_index, cell in enumerate(row, 1):
        if cell_index in (1, 2) and row_index not in (1, ):
            cell.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
        else:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

        cell.border = thin_border

        if row_index in (1, 2, 3):
            cell.font = Font(bold=True)

        if get_column_letter(cell_index) in ('I', 'K', 'M', 'O', 'Q', 'S') and row_index >= 3:
            cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
            if cell.value and '%' not in cell.value:
                cell_color = get_indicator_color(sheet.cell(row_index, cell_index).value, cell.value)
                cell.fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type='solid')
                cell.value = cell.value + '%'

for column_index, _ in enumerate(sheet.columns, 1):
    col_letter = get_column_letter(column_index)
    sheet.column_dimensions[col_letter].width = 11

sheet.column_dimensions["A"].width = 20
sheet.column_dimensions["B"].width = 30

for row_index, _ in enumerate(sheet.rows, 1):
    height = 18
    if row_index in (1, 2, 3):
        height = 30
    sheet.row_dimensions[row_index].height = height




book.save('test.xlsx')